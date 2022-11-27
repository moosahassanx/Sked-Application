# Imports
from msilib.schema import tables
import requests
from bs4 import BeautifulSoup
import pandas as pd
import xlrd
from openpyxl import Workbook, load_workbook

# Retrieving data
URL = "https://nccports.portauthoritynsw.com.au/eports/mobilemovements.asp"
page = requests.get(URL)
soup = BeautifulSoup(page.content, "html.parser")

# Setting out
vesselDates = soup.find_all('table', attrs={'border':'0'})
vesselValues = soup.find_all('table', attrs={'border':'1'})
vesselValues.pop(0)
vessels = []
vesselDateArr = []
vesselNames = []
tempVesselNames = []

# Parse dates
for i in vesselDates:
    date = i.find("b").string
    vessels.append({"date": date, "data": []})

# Parse data
for vesselIndex, j in enumerate(vesselValues):
    cells = j.findChildren("font")

    cellIndex = 0
    rowData = []
    for cell in cells:
        rowData.append(cell.string.strip())
        cellIndex += 1

        if(cellIndex == 5):
            vessels[vesselIndex]["data"].append({
                "Time": rowData[0],
                "From": rowData[1],
                "To": rowData[2],
                "Vessel": rowData[3],
                "Loa": rowData[4]
            })
            tempVesselNames.append(rowData[3])

            cellIndex = 0
            rowData = []

print("============= HTML PARSING =================")
print(vessels)
print(tempVesselNames)

# TODO: parse CSV data to draw comparisons between Mobile Movements and the BV Shipping List (ask rizwan bhai how it works)
print("============= SHEET READING =================")
shippingSheet = pd.read_excel("BV Shipping List 2022 - Email.xlsm", sheet_name=None)
shippingSheetName, df = next(iter(shippingSheet.items()))
df.columns = df.iloc[1]
df = df[2:]

# column cleanup
df = df[df['VESSEL'].notna()]

colVessel = df["VESSEL"]
for vessel in colVessel:
    print(vessel)

# Comparison between vessels and colVessel
for day in vessels:
    for vesselName in day['data']:
        if(vesselName['Vessel'] in tempVesselNames):
            vesselNames.append(vesselName['Vessel'])


# TODO: create new xlsm as target
print("============= SHEET CREATED =================")
wb = Workbook()
ws = wb.active
ws['A1'] = 42
ws.append(vesselNames)
wb.save('new_document.xlsm')
wb = load_workbook('new_document.xlsm', keep_vba=True)
wb.save('new_document.xlsm')

# TODO: brute testing + error catching + deployment for usage